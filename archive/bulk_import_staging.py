"""
bulk_import_staging.py — In-memory staging table for bulk CSV/Excel import.

Provides:
  - StagingTable: In-memory row store with auto-mapped columns
  - import_csv(path): Load CSV into a StagingTable
  - import_excel(path, sheet=None): Lazy-load Excel worksheet into StagingTable
  - auto_map_to_products(rows, mapping): Convert staged rows to product dicts

Integrates with:
  - database.add_product, database.update_product_full
  - openpyxl (for .xlsx lazy loading)
  - csv (for .csv)
"""
import csv
import os
import logging

import openpyxl

try:
    import native_accel
    _HAS_NATIVE_ACCEL = True
except ImportError:
    native_accel = None
    _HAS_NATIVE_ACCEL = False

log = logging.getLogger("bulk_import_staging")


class StagingTable:
    """In-memory staging table for bulk import rows.

    Columns are auto-discovered from the first CSV row or Excel header.
    Rows are stored as dicts keyed by column name.
    """

    def __init__(self, columns: list[str] = None, source_name: str = ""):
        self._columns: list[str] = columns or []
        self._rows: list[dict] = []
        self._source_name: str = source_name
        self._column_map: dict[str, str] = {}

    @property
    def columns(self) -> list[str]:
        return list(self._columns)

    @property
    def rows(self) -> list[dict]:
        return list(self._rows)

    @property
    def row_count(self) -> int:
        return len(self._rows)

    @property
    def source_name(self) -> str:
        return self._source_name

    def set_columns(self, columns: list[str]):
        """Set the column names (typically from header row)."""
        self._columns = list(columns)
        for col in self._columns:
            if col not in self._column_map:
                self._column_map[col] = ""

    def add_row(self, values: list | dict):
        """Add a row. If list, maps positionally to columns; if dict, uses keys."""
        if isinstance(values, dict):
            row = {col: values.get(col, values.get(self._column_map.get(col, col), "")) for col in self._columns}
        else:
            row = {}
            for i, col in enumerate(self._columns):
                row[col] = values[i] if i < len(values) else ""
        self._rows.append(row)

    def get_column(self, col_name: str) -> list:
        """Return all values from a single column."""
        return [row.get(col_name, "") for row in self._rows]

    def auto_map_csv_headers(self) -> dict[str, str]:
        """Attempt to map CSV headers to known product fields.

        Returns a dict of {csv_header: known_field}.

        Uses rapidfuzz (via native_accel) for fuzzy header matching when
        available, falling back to the existing 8-pass exact/normalized/
        substring algorithm.
        """
        known_fields = {
            "name": {"name", "product_name", "drug_name", "product", "item_name", "description"},
            "price": {"price", "unit_price", "cost", "retail_price", "selling_price"},
            "manufacturer_barcode": {"manufacturer_barcode", "mfg_barcode", "upc", "barcode", "mfg_code"},
            "internal_unique_barcode": {"internal_unique_barcode", "internal_barcode", "sku", "item_number", "id"},
            "dea_schedule": {"dea_schedule", "schedule", "dea", "drug_schedule"},
            "wholesale_price": {"wholesale_price", "wholesale", "wac", "awp"},
            "reorder_threshold": {"reorder_threshold", "min_stock", "reorder", "min_qty"},
            "expiry_date": {"expiry_date", "expiration_date", "exp_date", "expires"},
            "manufacture_date": {"manufacture_date", "mfg_date", "manufactured"},
            "vendor_name": {"vendor_name", "vendor", "supplier"},
            "status": {"status", "availability", "stock_status"},
        }

        if _HAS_NATIVE_ACCEL:
            try:
                result = native_accel.fuzzy_match_headers(self._columns, known_fields)
                if result:
                    self._column_map = result
                    return self._column_map
            except Exception as exc:
                log.debug("native_accel header matching failed, using fallback: %s", exc)

        self._column_map = {}
        for col in self._columns:
            lower = col.strip().lower().replace(" ", "_").replace("-", "_")
            compact = lower.replace("_", "")
            matched = False
            # 1. Exact field name match
            for field in known_fields:
                if lower == field:
                    self._column_map[col] = field
                    matched = True
                    break
            if matched:
                continue
            # 2. Exact alias match (normalized)
            for field, aliases in known_fields.items():
                for alias in aliases:
                    if lower == alias or compact == alias.replace("_", ""):
                        self._column_map[col] = field
                        matched = True
                        break
                if matched:
                    break
            if matched:
                continue
            # 3. Substring match (last resort, prefer longer aliases)
            for field, aliases in known_fields.items():
                for alias in sorted(aliases, key=len, reverse=True):
                    if alias in lower or lower.replace("_", "") in alias.replace("_", ""):
                        self._column_map[col] = field
                        matched = True
                        break
                if matched:
                    break
        return self._column_map

    def preview_rows(self, limit: int = 5) -> list[dict]:
        """Return the first `limit` rows for preview."""
        return self._rows[:limit]

    def clear(self):
        self._rows.clear()
        self._columns.clear()
        self._column_map.clear()

    def to_product_dicts(self) -> list[dict]:
        """Convert staged rows to product dicts using auto-mapped columns."""
        if not self._column_map:
            self.auto_map_csv_headers()

        products = []
        for row in self._rows:
            product = {}
            for csv_col, field_name in self._column_map.items():
                val = row.get(csv_col, "")
                if val == "":
                    continue
                if field_name in ("price", "wholesale_price"):
                    try:
                        product[field_name] = float(val)
                    except (ValueError, TypeError):
                        product[field_name] = 0.0
                elif field_name == "reorder_threshold":
                    try:
                        product[field_name] = int(float(val))
                    except (ValueError, TypeError):
                        product[field_name] = 0
                else:
                    product[field_name] = val
            products.append(product)
        return products


def import_csv(path: str) -> StagingTable:
    """Load a CSV file into a StagingTable.

    Uses csv.DictReader for header detection.
    """
    table = StagingTable(source_name=os.path.basename(path))
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            log.warning("CSV file is empty: %s", path)
            return table
        table.set_columns([h.strip() for h in headers])
        for raw_row in reader:
            if not raw_row or all(v.strip() == "" for v in raw_row):
                continue
            table.add_row(raw_row)
    table.auto_map_csv_headers()
    log.info("CSV import: %d rows, %d columns from %s", table.row_count, len(table.columns), path)
    return table


def import_excel(path: str, sheet: str = None) -> StagingTable:
    """Load an Excel (.xlsx) worksheet into a StagingTable.

    Uses openpyxl in read-only mode for memory efficiency on large files.
    If sheet is None, uses the first worksheet.
    """
    table = StagingTable(source_name=os.path.basename(path))

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        log.error("Excel import error: %s", e)
        raise

    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    # Read header row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value else "")

    table.set_columns(headers)

    # Read data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        values = [str(v) if v is not None else "" for v in row]
        table.add_row(values)

    wb.close()
    table.auto_map_csv_headers()
    log.info("Excel import: %d rows, %d columns from %s (sheet: %s)",
             table.row_count, len(table.columns), path, ws.title)
    return table


def commit_staged_products(table: StagingTable) -> dict:
    """Import all staged rows into the pharmacy database.

    Returns a dict with 'added', 'updated', 'errors' counts.
    """
    import database

    products = table.to_product_dicts()
    added = 0
    errors = 0

    for product in products:
        try:
            name = product.get("name", "")
            price = product.get("price", 0.0)
            mfg_barcode = product.get("manufacturer_barcode", "")
            internal_barcode = product.get("internal_unique_barcode", "")

            if not internal_barcode:
                errors += 1
                continue

            # Check if product already exists
            existing = database.get_product_by_internal_barcode(internal_barcode)
            if existing:
                database.update_product_full(
                    product_id=existing[0] if isinstance(existing, (list, tuple)) else existing.id,
                    name=name,
                    price=price,
                    manufacturer_barcode=mfg_barcode,
                    internal_barcode=internal_barcode,
                    expiry_date=product.get("expiry_date", ""),
                    manufacture_date=product.get("manufacture_date", ""),
                    status=product.get("status", "In Stock"),
                    vendor_name=product.get("vendor_name", "N/A"),
                    dea_schedule=product.get("dea_schedule", "OTC"),
                    wholesale_price=product.get("wholesale_price", 0.0),
                    reorder_threshold=product.get("reorder_threshold", 0),
                )
            else:
                database.add_product(
                    name=name,
                    price=price,
                    manufacturer_barcode=mfg_barcode,
                    internal_unique_barcode=internal_barcode,
                    expiry_date=product.get("expiry_date", ""),
                    manufacture_date=product.get("manufacture_date", ""),
                    vendor_name=product.get("vendor_name", "N/A"),
                    dea_schedule=product.get("dea_schedule", "OTC"),
                    wholesale_price=product.get("wholesale_price", 0.0),
                    reorder_threshold=product.get("reorder_threshold", 0),
                )
                added += 1
        except Exception as e:
            log.error("Failed to import staged product: %s", e)
            errors += 1

    return {"added": added, "updated": len(products) - added - errors, "errors": errors}
