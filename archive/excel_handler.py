import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import database
import barcode_logic
from async_ui import AsyncUI

try:
    import native_accel
    _HAS_NATIVE_ACCEL = True
except ImportError:
    native_accel = None
    _HAS_NATIVE_ACCEL = False


HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Database fields the import can populate (required fields highlighted)
DB_FIELDS = {
    "name":          {"label": "Product Name", "required": True,  "default": ""},
    "price":         {"label": "Price",        "required": True,  "default": ""},
    "expiry_date":   {"label": "Expiry Date",  "required": False, "default": ""},
    "mfg_barcode":   {"label": "Mfg Barcode",  "required": False, "default": ""},
    "vendor_name":   {"label": "Vendor",        "required": False, "default": "N/A"},
}

# Heuristic aliases: lowercased header -> db field key
HEADER_ALIASES = {
    "name":         "name",
    "product":      "name",
    "drug":         "name",
    "drug name":    "name",
    "product name": "name",
    "item":         "name",
    "item name":    "name",
    "price":        "price",
    "unit price":   "price",
    "cost":         "price",
    "expiry":       "expiry_date",
    "exp":          "expiry_date",
    "expir":        "expiry_date",
    "expiration":   "expiry_date",
    "expiry date":  "expiry_date",
    "sku":          "mfg_barcode",
    "barcode":      "mfg_barcode",
    "mfg barcode":  "mfg_barcode",
    "mfg":          "mfg_barcode",
    "upc":          "mfg_barcode",
    "vendor":       "vendor_name",
    "supplier":     "vendor_name",
    "vendor name":  "vendor_name",
    "supplier name":"vendor_name",
}


def read_excel_headers(file_path: str):
    """Read the first row (headers) from an Excel file.
    Returns (headers: list[str], row_count: int) or raises on error.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("No active sheet found in the Excel file.")
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = str(cell.value).strip() if cell.value is not None else ""
        headers.append(val)
    # Count data rows
    row_count = 0
    for _ in ws.iter_rows(min_row=2):
        row_count += 1
    wb.close()
    return headers, row_count


def auto_map_headers(excel_headers: list[str]):
    """Attempt to auto-match Excel headers to DB fields using aliases.
    Returns: {db_field_key: excel_col_index} for matched fields.
    Also returns unmatched header indices.
    """
    mapping = {}
    unmatched = []
    for idx, header in enumerate(excel_headers):
        key = header.lower().strip()
        if key in HEADER_ALIASES:
            db_field = HEADER_ALIASES[key]
            if db_field not in mapping:
                mapping[db_field] = idx
        else:
            unmatched.append((idx, header))
    return mapping, unmatched


def execute_import(file_path: str, column_map: dict, custom_field_map: dict = None,
                   default_values: dict = None, on_complete=None):
    """Import products from Excel using a validated column_map.
    column_map: {db_field_key: excel_col_index} — which Excel column maps to which DB field.
    custom_field_map: {excel_col_index: field_name} — extra columns to store as patient-style custom fields (future use).
    default_values: {db_field_key: default_value} — fallback values for DB fields not mapped.
    Runs in a background thread. Calls on_complete(imported_count, errors) when done.
    """
    default_values = default_values or {}

    def _worker():
        imported = 0
        errors = []
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return (0, ["No active sheet found."])

            # First pass: collect valid rows (some may be skipped)
            valid_rows: list[dict] = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    def cell_val(field_key):
                        col_idx = column_map.get(field_key)
                        if col_idx is None or col_idx >= len(row):
                            return default_values.get(field_key, "")
                        raw = row[col_idx]
                        return str(raw).strip() if raw is not None else ""

                    name = cell_val("name")
                    if not name:
                        continue

                    price_str = cell_val("price")
                    try:
                        price = float(price_str)
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num}: Invalid price '{price_str}'")
                        continue

                    expiry = cell_val("expiry_date")
                    sku = cell_val("mfg_barcode")
                    vendor = cell_val("vendor_name") or "N/A"

                    valid_rows.append({
                        "row_num": row_num,
                        "name": name,
                        "price": price,
                        "expiry": expiry,
                        "sku": sku,
                        "vendor": vendor,
                    })
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

            # Pre-generate all barcodes in batch (grouped by vendor)
            _vendor_counts: dict[str, int] = {}
            for vr in valid_rows:
                v = vr["vendor"]
                _vendor_counts[v] = _vendor_counts.get(v, 0) + 1

            _vendor_barcodes: dict[str, list[str]] = {}
            for vendor, count in _vendor_counts.items():
                if _HAS_NATIVE_ACCEL:
                    _vendor_barcodes[vendor] = native_accel.generate_batch_barcodes(vendor, count)
                else:
                    _vendor_barcodes[vendor] = [barcode_logic.generate_internal_barcode(vendor) for _ in range(count)]

            # Assign barcodes from pre-generated batches
            _vendor_idx: dict[str, int] = {v: 0 for v in _vendor_counts}
            for vr in valid_rows:
                v = vr["vendor"]
                vr["barcode"] = _vendor_barcodes[v][_vendor_idx[v]]
                _vendor_idx[v] += 1

            # Second pass: insert into database
            for vr in valid_rows:
                database.add_product(
                    name=vr["name"],
                    price=vr["price"],
                    manufacturer_barcode=vr["sku"],
                    internal_unique_barcode=vr["barcode"],
                    expiry_date=vr["expiry"],
                    manufacture_date="",
                    vendor_name=vr["vendor"],
                )
                imported += 1

            wb.close()
        except Exception as e:
            errors.append(f"File error: {str(e)}")

        return (imported, errors)

    def _on_done(result, error=None):
        if on_complete:
            if isinstance(result, tuple):
                on_complete(result[0], result[1])

    return AsyncUI.get().run(_worker, callback=_on_done)


# ── Export Functions ──────────────────────────────────────────────────────────

def export_to_excel(data_list: list, headers: list, output_path: str, on_complete=None):
    """Export data to a formatted Excel file."""
    def _worker():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        for row_idx, row_data in enumerate(data_list, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")

        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row_idx in range(2, len(data_list) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 40)

        wb.save(output_path)
        wb.close()
        return output_path

    def _on_done(result, error=None):
        if on_complete:
            if error or result is None:
                on_complete(None)
            else:
                on_complete(result)

    return AsyncUI.get().run(_worker, callback=_on_done)


def export_inventory(output_path: str, on_complete=None):
    """Export current in-stock inventory to Excel."""
    batches = database.get_all_in_stock_batches()
    headers = ["Name", "Price", "Int. Barcode", "Vendor", "Expiry", "Mfg Date", "Mfg Barcode"]
    data = []
    for b in batches:
        _, name, price, mfg_barcode, int_barcode, status, expiry, mfg_date, vendor = b
        data.append((name, price, int_barcode, vendor or "N/A", expiry or "", mfg_date or "", mfg_barcode))
    return export_to_excel(data, headers, output_path, on_complete=on_complete)
